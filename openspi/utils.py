import os
import openpyxl as xl
import pandas as pd

#from .metadata import _xlsx_metadata

def reformat_path(path):
    """
    Reformat a raw string file path (path\\\\to\\\\folder) to contain double back-
    slashes (path\\\\\\\\to\\\\\\\\folder), as is required by the R programming language.
    Doesn't do anything if the path uses forward slashes (/)

    Parameters
    ----------
    path : str
        The path to be reformatted.

    Returns
    -------
    reformatted_path : str
        All \\ are replaced with \\\\.

    """

    reformatted_path = path.replace("\\", "\\\\")
    return reformatted_path


def identical_list_items(lst):
    """
    Checks if all the values in a given list are the same.

    Parameters
    ----------
    lst : list
        A list of values.

    Returns
    -------
    bool
        Returns ``True`` if all the values are the same, ``False`` if not.

    """
    return len(set(lst)) == 1


def plastic_matches_checked(row, index):
    """
    Appends a note to a given list which marks the index of the subsequent
    plastic match.

    Parameters
    ----------
    row : list
        A list containing data for one library match for one file.
    index : int
        The index of the subsequent plastic match.
        (see ``subsequent_matches_checked``)

    Returns
    -------
    row : list
        An updated version of the original ``row`` var. It now contains an
        additional item.

    """

    if index == 1:
        row.append("second match")

    elif index == 2:
        row.append("third match")

    elif index == 3:
        row.append("fourth match")

    elif index == 4:
        row.append("fifth match")

    return row


def nonpolymer_matches_checked(row, nested_list):
    """
    Checks a given dataframe (in the form of a nested list) to determine if all
    matches are empty wells or not.

    Parameters
    ----------
    row : list
        A list containing data for one library match for one file.
    nested_list : list
        A list of lists, where each list in the outer list represents a row.

    Returns
    -------
    row : list
        An updated version of the original``row`` var. It now contains an
        additional item.

    """

    if "empty well" in nested_list:
        # Check to see if 'empty well' is the only value in that list
        if identical_list_items(nested_list) == True:
            row.append("all top matches empty wells")
        else:
            row.append("all top matches nonpolymer")
    else:
        row.append("all top matches nonpolymer")

    return row


def subsequent_matches_checked(df, nested_list):
    """
    Checks if the first match (highest r val) for each sample is plastic, and
    if it is not, it checks the subsequent matches for any plastic matches, and
    if there are, it makes a note of its place (1st, 2nd, etc). If there are no
    plastic matches, it will check if all subsequent matches are for an empty
    well and make a note. If that is not the case, it will note that all
    subsequent matches are 'nonpolymer.'

    Parameters
    ----------
    df : df
        A Pandas dataframe.
    nested_list : list
        A list of lists, where each list in the outer list represents a row.

    Returns
    -------
    nested_list : list
        An updated version of the original``nested_list``. Each inner list is
        now one item longer.

    """

    # Send plastic classification column to list
    polymer = df["plastic_or_not"].values.tolist()

    # Proceed if the first match is not plastic
    if polymer[0] == "not plastic":
        if "plastic" in polymer:

            # Identify index of first occurrence of 'plastic' and pull the
            # corresponding row to be added to the updated summary sheet
            index = polymer.index("plastic")
            row = df.iloc[index].tolist()

            # Add a note to the row indicating its match # (1st, 2nd, 3rd, etc)
            row = plastic_matches_checked(row, index)

        else:
            # Send the spectrum_identity column to a list
            material = df["spectrum_identity"].values.tolist()

            # Send the first row to a list and add a note indicating if all
            # matches are empty wells or simply nonpolymer
            row = df.iloc[0].tolist()
            row = nonpolymer_matches_checked(row, material)

    else:
        # If the first match is polymer, send the first row to a list
        row = df.iloc[0].tolist()
        row.append(" ")

    # Add row to nested_list
    nested_list.append(row)

    return nested_list


def save_df_to_excel(excel_path, df, sheetname):
    """
    Saves a Pandas dataframe to an Excel file. If the Excel file does not yet
    exist, it will create it and save the sheet.

    Parameters
    ----------
    excel_path : str
        The full path to an .xlsx file.
    df : df
        The dataframe to be saved as an Excel sheet.
    sheetname : str
        The name of the Excel sheet.

    Returns
    -------
    None.

    """

    if os.path.exists(excel_path) == False:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheetname, index=False)

    else:
        with pd.ExcelWriter(
            excel_path, engine="openpyxl", if_sheet_exists="replace", mode="a"
        ) as writer:
            df.to_excel(writer, sheet_name=sheetname, index=False)

    _xlsx_metadata(excel_path)


def check_excel_sheet(excel_path, sheetname):
    """
    Checks if an Excel spreadsheet exists within the (already existing) Excel
    workbook and creates one if not. Note that this differs from
    ``save_df_to_excel`` and is used for directly editing cells in an Excel
    spreadsheet.

    Parameters
    ----------
    excel_path : str
        The full path to an already existing .xlsx file.
    sheetname : str
        The name of the Excel sheet.

    Returns
    -------
    ws : openpyxl worksheet object
        An Excel sheet with the desired ``sheetname``.

    """

    # Check if the designated sheetname exists
    if str(sheetname) in excel_path.sheetnames:
        # If the sheetname exists, assign it to ws
        ws = excel_path[str(sheetname)]
    else:
        # If the sheetname does not exist, create it and assign to ws
        ws = excel_path.create_sheet(str(sheetname))

    return ws


def empty_wells_count(df):
    """
    Counts how many times 'empty well' appears in a given dataframe.

    Parameters
    ----------
    df : df
        A Pandas dataframe.

    Returns
    -------
    count : int
        The number of times ``'empty well'`` appears in the
        ``'spectrum_identity'`` column.

    """

    # Send all rows that contain 'empty well' in the 'spectrum_identity' column
    # to a new dataframe
    df_empty_wells = df[df["spectrum_identity"] == "empty well"]

    # Count the length of the dataframe
    count = len(df_empty_wells)

    return count


def notes_sheet(excel_path):
    """
    Adds a 'Notes' sheet with the number of nonpolymer matches and empty wells

    Parameters
    ----------
    excel_path : str
        The full path to an .xlsx file.

    Returns
    -------
    None.

    """

    # Open the Excel workbook
    wb1 = xl.load_workbook(excel_path)

    # Check if 'Notes' sheet exists and create one if not
    notes_sheet = check_excel_sheet(wb1, "Notes")

    # Open the Summary sheet into a pandas dataframe
    xlsx_file = pd.ExcelFile(excel_path)
    df_summary = pd.read_excel(xlsx_file, "Summary")

    # Send all rows with not plastic matches to a dataframe
    df_not_plastic = df_summary[df_summary["plastic_or_not"] == "not plastic"]

    # Check if the 'not plastic' dataframe is emtpy
    if df_not_plastic.empty:
        print("No nonplastic matches.")

    else:
        # Count how many empty well matches are in the dataframe
        empty_wells = empty_wells_count(df_not_plastic)

        # Open the Updated Summary sheet into a pandas dataframe
        df_updated_summary = pd.read_excel(xlsx_file, "Updated Summary")

        # Send all rows containing the matching text to a dataframe and count
        # the length of the dataframe
        df_all_empty = df_updated_summary[
            df_updated_summary["matches_checked"] == "all top matches empty wells"
        ]
        totally_empty_count = len(df_all_empty)

        # Compare the empty well count to the nonpolymer matches count
        empty_matches = (
            str(empty_wells)
            + " out of "
            + str(len(df_not_plastic))
            + " nonplastic matches are empty wells."
        )

        notes_sheet["A3"] = empty_matches

        # Compare the totally empty well count (5/5 empty) to the empty well
        # count (first match empty)
        empty = (
            str(totally_empty_count)
            + " out of "
            + str(empty_wells)
            + " empty well matches have 5/5 top hits for empty wells."
        )

        notes_sheet["A4"] = empty

        wb1.save(str(excel_path))

        print("Workbook saved to " + excel_path)

    _xlsx_metadata(excel_path)


def list_to_df_to_sheet(df_lst, columns_list, excel_path, sheet_name):
    """
    Converts a nested list into a Pandas dataframe, which is then saved to an
    Excel worksheet.

    Parameters
    ----------
    excel_path : str
        The full path to an .xlsx file.
    df_lst : list
        The nested list.
    columns_list : list
        The list of column names.
    sheet_name : str
        The name of the Excel sheet.

    Returns
    -------
    None.

    """

    df = pd.DataFrame(df_lst, columns=columns_list)
    save_df_to_excel(excel_path, df, sheet_name)